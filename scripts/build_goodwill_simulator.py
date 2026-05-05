"""
영업권 평가 Excel 시뮬레이터 빌더
산출물: outputs/영업권평가시뮬레이터_TEMPLATE.xlsx

시트 구성:
  1. 입력값      — 재무제표·환원율·자본환원율
  2. 초과이익환원법 — 상증령§59 (법인전환·가업승계·상속증여 표준)
  3. 수익환원법    — DCF 5~10년 예측 + 잔존가치
  4. 거래사례비교법 — EV/EBITDA · P/E 멀티플
  5. 잔여접근법    — 총사업가치 - 유형순자산
  6. 비교표        — 4가지 평가법 결과 + 추천
  7. 4자관점매트릭스 — 4자관점 × 3시점 12셀
  8. 시나리오      — 보수·중립·낙관 3종

실행: python scripts/build_goodwill_simulator.py
요구: openpyxl (pip install openpyxl)
"""
from __future__ import annotations
from pathlib import Path
from datetime import date
import sys


def _try_import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference
        return openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, BarChart, Reference
    except ImportError:
        return None


def _apply_header(ws, row, col, text, bg_color="1F4E79", font_color="FFFFFF", bold=True, Font=None, PatternFill=None, Alignment=None):
    cell = ws.cell(row=row, column=col, value=text)
    if Font:
        cell.font = Font(bold=bold, color=font_color, size=10)
    if PatternFill:
        cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
    if Alignment:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def build_goodwill_simulator(output_path: str | None = None) -> str:
    """Excel 시뮬레이터 생성"""
    result = _try_import_openpyxl()
    if result is None:
        return _build_csv_fallback(output_path)

    openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter, BarChart, Reference = result

    wb = openpyxl.Workbook()

    # ── 색상 상수 ─────────────────────────────────────────────────────────────
    NAVY   = "1F4E79"
    GOLD   = "C00000"
    LIGHT  = "D6E4F0"
    GREEN  = "375623"
    ORANGE = "E26B0A"
    GRAY   = "595959"

    def hdr(ws, row, col, text, bg=NAVY, fg="FFFFFF"):
        return _apply_header(ws, row, col, text, bg, fg, True, Font, PatternFill, Alignment)

    def num_cell(ws, row, col, val, fmt="#,##0"):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="right")
        return c

    def pct_cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = "0.0%"
        c.alignment = Alignment(horizontal="right")
        return c

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 1: 입력값
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "①입력값"
    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 18

    ws1.merge_cells("A1:C1")
    hdr(ws1, 1, 1, "영업권 평가 — 입력 파라미터 (상증령§59 기준)")
    ws1.row_dimensions[1].height = 24

    ws1.cell(row=2, column=1, value="평가 기준일").font = Font(bold=True)
    ws1.cell(row=2, column=2, value=str(date.today()))
    ws1.cell(row=2, column=3, value="YYYY-MM-DD")

    ws1.cell(row=3, column=1, value="평가 목적").font = Font(bold=True)
    ws1.cell(row=3, column=2, value="법인전환")
    ws1.cell(row=3, column=3, value="법인전환|M&A|가업승계|상속증여")

    # 구분선
    hdr(ws1, 5, 1, "재무 데이터 (최근 3개년 평균)", LIGHT, GRAY)
    ws1.merge_cells("A5:C5")

    fin_rows = [
        ("최근 3년 평균 순손익 (원)", 500_000_000, "당기순이익 평균"),
        ("자기자본 시가 (원)", 1_500_000_000, "자산시가 - 부채시가"),
        ("매출액 (원)", 2_000_000_000, "최근 1개년"),
        ("EBITDA 추정 (원)", 600_000_000, "순손익 × 1.2 추정"),
        ("유형순자산 (원)", 1_000_000_000, "건물·기계·설비 시가"),
    ]
    for i, (label, val, note) in enumerate(fin_rows, start=6):
        ws1.cell(row=i, column=1, value=label)
        num_cell(ws1, i, 2, val)
        ws1.cell(row=i, column=3, value=note).font = Font(color=GRAY, italic=True, size=9)

    hdr(ws1, 12, 1, "평가 파라미터 (수정 가능)", LIGHT, GRAY)
    ws1.merge_cells("A12:C12")

    param_rows = [
        ("환원율 (초과이익환원법·§59)", 0.06, "일반: 5~7%"),
        ("자본환원율 (정상수익률·§54)", 0.10, "법령 고정: 10%"),
        ("할인율 WACC (수익환원법)", 0.06, "WACC: 5~8%"),
        ("연간 성장률 가정 (DCF)", 0.03, "3~5%"),
        ("EV/EBITDA 멀티플 (거래사례)", 5.0, "업종별 2~10배"),
        ("P/E 멀티플 (거래사례)", 10.0, "업종별 5~20배"),
        ("DCF 예측 기간 (수익환원법)", 5, "5~10년"),
    ]
    for i, (label, val, note) in enumerate(param_rows, start=13):
        ws1.cell(row=i, column=1, value=label)
        if isinstance(val, float) and val < 1:
            pct_cell(ws1, i, 2, val)
        else:
            num_cell(ws1, i, 2, val, "#,##0.00")
        ws1.cell(row=i, column=3, value=note).font = Font(color=GRAY, italic=True, size=9)

    ws1.cell(row=21, column=1, value="※ 법령 기준: 상증법§64·상증령§59·§54·법§52·소§94·조특§32")
    ws1.cell(row=21, column=1).font = Font(color=GOLD, size=9)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 2: 초과이익환원법 (상증령§59)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("②초과이익환원법")
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 22

    ws2.merge_cells("A1:B1")
    hdr(ws2, 1, 1, "초과이익환원법 (상증령§59) — 법인전환·가업승계·상속증여 표준")

    steps = [
        ("STEP 1. 정상 순이익 산정", None),
        ("  자기자본 시가 (원)", "=①입력값!B7"),
        ("  자본환원율 (§54)", "=①입력값!B14"),
        ("  정상 순이익 = 자기자본 × 자본환원율", "=B4*B5"),
        ("STEP 2. 초과이익 산정", None),
        ("  최근 3년 평균 순손익 (원)", "=①입력값!B6"),
        ("  초과이익 = 순손익 - 정상순이익 (원)", "=MAX(0,B8-B6)"),
        ("STEP 3. 환원계수 산정", None),
        ("  환원율", "=①입력값!B13"),
        ("  환원계수 = 1 / 환원율", "=IFERROR(1/B11,0)"),
        ("STEP 4. 영업권 산출", None),
        ("  ★ 영업권 = 초과이익 × 환원계수 (원)", "=B9*B12"),
        ("  5년 균등상각 (연 손금)", "=B14/5"),
        ("  양도소득세 추정 (소§94·22%)", "=B14*0.22"),
    ]
    for i, (label, val) in enumerate(steps, start=2):
        ws2.cell(row=i, column=1, value=label)
        if val:
            if str(val).startswith("="):
                ws2.cell(row=i, column=2, value=val).number_format = "#,##0.00"
            else:
                ws2.cell(row=i, column=2, value=val)
        if label.startswith("STEP") or label.startswith("  ★"):
            ws2.cell(row=i, column=1).font = Font(bold=True, color=NAVY if "STEP" in label else GOLD)

    ws2.cell(row=18, column=1, value="법령 근거: 상증령§59, §54 / 소득세법§94 ①4 / 조특§32 / 법§24·법시§35")
    ws2.cell(row=18, column=1).font = Font(color=GOLD, size=9)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 3: 수익환원법 (DCF)
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("③수익환원법DCF")
    ws3.column_dimensions["A"].width = 30
    for col in range(2, 12):
        ws3.column_dimensions[get_column_letter(col)].width = 14

    ws3.merge_cells("A1:K1")
    hdr(ws3, 1, 1, "수익환원법 (DCF) — M&A · 수익성 기업 평가")

    ws3.cell(row=2, column=1, value="구분")
    for yr in range(1, 6):
        ws3.cell(row=2, column=yr + 1, value=f"{yr}년차").font = Font(bold=True)
    ws3.cell(row=2, column=7, value="잔존가치").font = Font(bold=True, color=GOLD)
    ws3.cell(row=2, column=8, value="합계 영업권").font = Font(bold=True, color=GOLD)

    ws3.cell(row=3, column=1, value="초과이익 (성장 반영)")
    ws3.cell(row=4, column=1, value="현가계수 (1/(1+r)^n)")
    ws3.cell(row=5, column=1, value="현재가치")
    ws3.cell(row=6, column=1, value="잔존가치 (고든성장)").font = Font(italic=True)
    ws3.cell(row=7, column=1, value="★ DCF 영업권 합계").font = Font(bold=True, color=GOLD)

    ws3.cell(row=10, column=1, value="할인율 (WACC) 참조")
    ws3.cell(row=10, column=2, value="=①입력값!B15")
    ws3.cell(row=11, column=1, value="연간 성장률")
    ws3.cell(row=11, column=2, value="=①입력값!B16")
    ws3.cell(row=12, column=1, value="법령 근거: DCF + 상증령§59 병용")
    ws3.cell(row=12, column=1).font = Font(color=GOLD, size=9)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 4: 거래사례비교법
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("④거래사례비교법")
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 18

    ws4.merge_cells("A1:C1")
    hdr(ws4, 1, 1, "거래사례비교법 — 동종업종 EV/EBITDA · P/E 멀티플")

    comp_rows = [
        ("EV/EBITDA 방식", None),
        ("  EBITDA (원)", "=①입력값!B9"),
        ("  EV/EBITDA 멀티플", "=①입력값!B17"),
        ("  기업가치(EV) = EBITDA × 멀티플", "=B4*B5"),
        ("  유형순자산 (원)", "=①입력값!B10"),
        ("  영업권 = EV - 유형순자산", "=MAX(0,B6-B7)"),
        ("P/E 방식", None),
        ("  평균 순손익 (원)", "=①입력값!B6"),
        ("  P/E 멀티플", "=①입력값!B18"),
        ("  주식가치 = 순손익 × P/E", "=B10*B11"),
        ("  영업권 = 주식가치 - 유형순자산", "=MAX(0,B12-B7)"),
        ("★ 평균 영업권 = (EV/EBITDA + P/E) / 2", "=(B8+B13)/2"),
        ("법령 근거: 시가 참고 (법§52) + DART 동종업종 비교", None),
    ]
    for i, (label, val) in enumerate(comp_rows, start=2):
        ws4.cell(row=i, column=1, value=label)
        if val and str(val).startswith("="):
            ws4.cell(row=i, column=2, value=val).number_format = "#,##0.00"
        if label.startswith("STEP") or label.startswith("★") or (label and not label.startswith(" ")):
            ws4.cell(row=i, column=1).font = Font(bold=True, color=NAVY if "방식" in label else GOLD)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 5: 잔여접근법
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("⑤잔여접근법")
    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 20

    ws5.merge_cells("A1:B1")
    hdr(ws5, 1, 1, "잔여접근법 — 총사업가치 - 식별가능 유형순자산")

    residual_rows = [
        ("총사업가치 (수익환원법 활용·원)", "=③수익환원법DCF!B7"),
        ("식별가능 유형순자산 (원)", "=①입력값!B10"),
        ("★ 영업권 = 총사업가치 - 유형순자산 (원)", "=MAX(0,B3-B4)"),
        ("법령 근거: K-IFRS 1103 사업결합 + 법§52", ""),
    ]
    for i, (label, val) in enumerate(residual_rows, start=2):
        ws5.cell(row=i, column=1, value=label)
        if val and str(val).startswith("="):
            ws5.cell(row=i, column=2, value=val).number_format = "#,##0.00"
        if "★" in label:
            ws5.cell(row=i, column=1).font = Font(bold=True, color=GOLD)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 6: 비교표 + 추천
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("⑥비교표_추천")
    ws6.column_dimensions["A"].width = 22
    ws6.column_dimensions["B"].width = 20
    ws6.column_dimensions["C"].width = 16
    ws6.column_dimensions["D"].width = 30

    ws6.merge_cells("A1:D1")
    hdr(ws6, 1, 1, "4가지 평가법 비교 + 케이스별 추천")

    hdr(ws6, 2, 1, "평가법", LIGHT, GRAY)
    hdr(ws6, 2, 2, "영업권 산출액 (원)", LIGHT, GRAY)
    hdr(ws6, 2, 3, "법령 근거", LIGHT, GRAY)
    hdr(ws6, 2, 4, "적용 케이스", LIGHT, GRAY)

    methods = [
        ("초과이익환원법", "=②초과이익환원법!B14", "상증령§59·§54", "법인전환·가업승계·상속증여 ★표준"),
        ("수익환원법(DCF)", "=③수익환원법DCF!B7", "DCF·WACC", "M&A·투자 가치 산정"),
        ("거래사례비교법", "=④거래사례비교법!B14", "법§52 시가 참고", "분쟁조정·객관적 시장가"),
        ("잔여접근법", "=⑤잔여접근법!B4", "K-IFRS 1103", "총사업가치 기반 담보평가"),
    ]
    for i, (name, formula, law, case) in enumerate(methods, start=3):
        ws6.cell(row=i, column=1, value=name)
        ws6.cell(row=i, column=2, value=formula).number_format = "#,##0"
        ws6.cell(row=i, column=3, value=law)
        ws6.cell(row=i, column=4, value=case)

    ws6.cell(row=8, column=1, value="★ 케이스별 추천").font = Font(bold=True, color=GOLD)
    case_rec = [
        ("법인전환", "초과이익환원법", "세무서 인정도 최고 (상증령§59)"),
        ("가업승계·상속증여", "초과이익환원법", "상증법§64 강제 적용"),
        ("M&A·인수합병", "수익환원법+거래사례 평균", "수익 창출 능력 중심"),
        ("담보·신용평가", "잔여접근법", "유형자산 제외 무형가치"),
    ]
    for i, (case, method, note) in enumerate(case_rec, start=9):
        ws6.cell(row=i, column=1, value=case)
        ws6.cell(row=i, column=2, value=method).font = Font(bold=True, color=NAVY)
        ws6.cell(row=i, column=4, value=note)

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 7: 4자관점 × 3시점 12셀
    # ══════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("⑦4자관점매트릭스")
    ws7.merge_cells("A1:D1")
    hdr(ws7, 1, 1, "4자관점(법인·양도자·과세관청·금융기관) × 3시점(Pre·Now·Post) 12셀")

    hdr(ws7, 2, 1, "관점", LIGHT, GRAY)
    hdr(ws7, 2, 2, "사전 (Pre)", LIGHT, GRAY)
    hdr(ws7, 2, 3, "현재 (Now)", LIGHT, GRAY)
    hdr(ws7, 2, 4, "사후 (Post)", LIGHT, GRAY)

    matrix_data = [
        ("법인",
         "재무 데이터 정합·평가 기준일 확정",
         "영업권 자산 인식·취득가액 기장",
         "5년 균등상각 (연 손금)·손상 분기 점검"),
        ("양도자(개인)",
         "영업권 가액 협상·평가법 검토",
         "양도소득 확정·세금 납부 or 이월과세",
         "이월과세 5년 추적·경정청구 가능성"),
        ("과세관청",
         "상증령§59 산식 사전심사·시가 기준",
         "평가 적정성 검토·법§52 시가 검증",
         "세무조사·추징 가능성·경정청구 처리"),
        ("금융기관",
         "영업권 담보가치 인정 여부 협의",
         "영업권 반영 자기자본비율 재산정",
         "상각 후 재평가·여신 갱신"),
    ]
    for i, (party, pre, now, post) in enumerate(matrix_data, start=3):
        ws7.cell(row=i, column=1, value=party).font = Font(bold=True)
        ws7.cell(row=i, column=2, value=pre)
        ws7.cell(row=i, column=3, value=now)
        ws7.cell(row=i, column=4, value=post)
        for col in range(1, 5):
            ws7.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    for col in range(1, 5):
        ws7.column_dimensions[get_column_letter(col)].width = 28
        ws7.row_dimensions[i].height = 48

    # ══════════════════════════════════════════════════════════════════════════
    # 시트 8: 시나리오 3종
    # ══════════════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("⑧시나리오3종")
    ws8.column_dimensions["A"].width = 25
    ws8.column_dimensions["B"].width = 18
    ws8.column_dimensions["C"].width = 18
    ws8.column_dimensions["D"].width = 18

    ws8.merge_cells("A1:D1")
    hdr(ws8, 1, 1, "시나리오 3종 (초과이익환원법 기준 — 환원율·정상수익률 변동)")

    hdr(ws8, 2, 1, "항목", LIGHT, GRAY)
    hdr(ws8, 2, 2, "보수 시나리오", GOLD, "FFFFFF")
    hdr(ws8, 2, 3, "중립 시나리오", NAVY, "FFFFFF")
    hdr(ws8, 2, 4, "낙관 시나리오", GREEN, "FFFFFF")

    scenario_rows = [
        ("환원율", 0.07, 0.06, 0.05),
        ("정상수익률 (§54)", 0.12, 0.10, 0.08),
        ("자기자본 (원)", "=①입력값!B7", "=①입력값!B7", "=①입력값!B7"),
        ("정상순이익 (원)", "=D4*B4", "=D4*C4", "=D4*D4"),
        ("평균순손익 (원)", "=①입력값!B6", "=①입력값!B6", "=①입력값!B6"),
        ("초과이익 (원)", "=MAX(0,B7-B6)", "=MAX(0,C7-C6)", "=MAX(0,D7-D6)"),
        ("환원계수", "=IFERROR(1/B3,0)", "=IFERROR(1/C3,0)", "=IFERROR(1/D3,0)"),
        ("★ 영업권 (원)", "=B8*B9", "=C8*C9", "=D8*D9"),
        ("5년 상각 (연, 원)", "=B10/5", "=C10/5", "=D10/5"),
        ("양도소득세 추정 (원)", "=B10*0.22", "=C10*0.22", "=D10*0.22"),
    ]
    for i, row_data in enumerate(scenario_rows, start=3):
        label = row_data[0]
        ws8.cell(row=i, column=1, value=label)
        if "★" in str(label):
            ws8.cell(row=i, column=1).font = Font(bold=True, color=GOLD)
        for col, val in enumerate(row_data[1:], start=2):
            c = ws8.cell(row=i, column=col, value=val)
            if isinstance(val, float) and val < 1:
                c.number_format = "0.0%"
            elif str(val).startswith("="):
                c.number_format = "#,##0"
            c.alignment = Alignment(horizontal="right")

    ws8.cell(row=15, column=1, value="법령: 상증령§59·§54 / 소§94 / 조특§32 / 법§24·법시§35")
    ws8.cell(row=15, column=1).font = Font(color=GOLD, size=9)

    # ── 저장 ─────────────────────────────────────────────────────────────────
    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    out_path = output_path or str(out_dir / "영업권평가시뮬레이터_TEMPLATE.xlsx")
    wb.save(out_path)
    return out_path


def _build_csv_fallback(output_path: str | None = None) -> str:
    """openpyxl 없을 때 CSV 대체 파일 생성"""
    out_dir = Path("outputs")
    out_dir.mkdir(exist_ok=True)
    csv_path = output_path or str(out_dir / "영업권평가시뮬레이터_TEMPLATE.csv")

    lines = [
        "영업권 평가 시뮬레이터 (openpyxl 설치 후 Excel 버전 생성 가능)",
        "pip install openpyxl",
        "",
        "=== 초과이익환원법 (상증령§59) ===",
        "항목,값,설명",
        "최근3년평균순손익,500000000,원",
        "자기자본시가,1500000000,원",
        "자본환원율,0.10,§54 기준",
        "환원율,0.06,5~7%",
        "정상순이익,150000000,자기자본×10%",
        "초과이익,350000000,순손익-정상순이익",
        "환원계수,16.67,1/환원율",
        "영업권,5833333333,초과이익×환원계수",
        "",
        "=== 법령 ===",
        "상증법§64 / 상증령§59·§54 / 법§52·§24·법시§35 / 소§94 / 조특§32",
    ]
    Path(csv_path).write_text("\n".join(lines), encoding="utf-8-sig")
    return csv_path


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else None
    out_path = build_goodwill_simulator(output)
    print("OK 영업권 평가 Excel 시뮬레이터 생성: " + out_path)
